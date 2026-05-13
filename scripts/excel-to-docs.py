#!/usr/bin/env python3
"""Convert OneStore xlsx (+ a small manual config) → notion-documents.json.

Run this on the personal laptop whenever the xlsx is refreshed or you want
to update the Important/Roadmap flags. It is NOT run by GitHub Actions.

Usage:
    python scripts/excel-to-docs.py \
        --xlsx "/path/to/20260512-list-onestore.xlsx" \
        --extras scripts/manual-docs.json \
        --out notion-documents.json
"""

import argparse, csv, json, os, sys
from datetime import datetime, timezone, timedelta

# ----- Portfolio normalization -----
NORMALIZE = {'OWM': 'OMW', 'XTG': 'XTM'}
VALID_PORTFOLIO = {
    'FUW', 'Fiber-Sensing', 'FlexILS', 'GX', 'ICE-X', 'IPM', 'LX', 'OMW',
    'PSD', 'PSI-L', 'PSI-M', 'PSS', 'PSS-HC', 'PSS-X', 'QSN', 'Service',
    'Wavesuite', 'XTM',
}

# ----- Tag heuristics (keyword → tag) -----
TAG_RULES = [
    ('Visio Stencil',     ['visio','.vss','.vsdx','stencil','icons','rackface','link diagram']),
    ('Demo Video',        ['demo video','.mp4','animation','demonstration video',' demo ']),
    ('Webinar',           ['webinar']),
    ('Datasheet',         ['datasheet','data sheet']),
    ('Whitepaper',        ['whitepaper','white paper']),
    ('Application Note',  ['application note','app note']),
    ('Case Study',        ['case study','success stor','customer momentum','customer briefs','newton win','modernizes']),
    ('GRTS',              ['grts','get ready to sell']),
    ('Release Notes',     ['new in 1830','new in r','whats new',"what's new"]),
    ('Pricing & Licensing',['pricing','licens','subscription','eula','license capabilities']),
    ('Competitive Intel', ['competitive','competitor','competition','countering','vs ciena','vs huawei']),
    ('Exec Overview',     ['exec overview','executive overview','exec brief','executive brief','executive presentation','business vision']),
    ('Deep Dive',         ['deep dive','deep-dive','technical details','tech details','tech overview','technical overview','fundamentals','tutorial','understanding']),
    ('Sales Presentation',['sales presentation','sales training','partner training','customer presentation','training']),
    ('Roadmap',           ['roadmap','launch']),
    ('Customer-facing',   ['customer-facing','customer facing','customer presentation','customer overview']),
    ('Internal',          ['internal','rbc','pre-ga','pre ga']),
    ('Partner',           ['partner','partners']),
    ('DCI',               ['dci','data center interconnect','data centre interconnect']),
    ('Subsea',            ['subsea','slte','sub-sea','submarine','festoon']),
    ('Long Haul',         ['long haul','long-haul']),
    ('Metro',             ['metro','metro edge','metro access']),
    ('Mission Critical',  ['mission critical','mission-critical','enterprise','utilities','railway','transportation','airport','healthcare']),
    ('Quantum-Safe',      ['quantum','qsn','quantum-safe','quantum safe','pqc','qkd','encryption']),
    ('Fiber Sensing',     ['fiber sensing','fiber-sensing',' sensing','otdr']),
    ('5G / Mobile',       ['5g','mobile fronthaul','fronthaul','mobile transport','o-ran','oran','anyhaul','mobile backhaul']),
    ('Automation',        ['automation','automate','wavesuite','closed loop','workflow']),
    ('Pluggables',        ['pluggable','400zr','800zr','zr+','zr/zr+','ice-x','multihaul','router-pluggable','dco','coherent pluggable']),
    ('AI / Data Center',  [' ai ','ai era','data center','hyperscale','data centre']),
    ('Sustainability',    ['sustainab','esg','energy','green','power-efficient']),
    ('OTN',               ['otn','g.709','fgotn']),
    ('Sync / Timing',     ['sync','synchronization','timing','ptp','1588']),
    ('IP-Optical',        ['ip-optical','ip optical','coherent routing','cross-domain','cross domain','nsp']),
    ('ROADM',             ['roadm','iroadm','wss','cdc-f']),
]


def assign_tags(name, desc):
    text = f"{name} {desc}".lower()
    tags = []
    for tag, kws in TAG_RULES:
        for kw in kws:
            if kw.lower() in text:
                if tag not in tags:
                    tags.append(tag)
                break
    if not tags:
        tags.append('Sales Presentation')
    return tags


def s(v):
    return (str(v).strip() if v else '')


def read_xlsx(xlsx_path):
    try:
        import openpyxl
    except ImportError:
        sys.exit('ERROR: openpyxl not installed. Run: pip install openpyxl')
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['query (2)'] if 'query (2)' in wb.sheetnames else wb.active
    rows = []
    for row in ws.iter_rows(min_row=2):
        # Columns: A=ID, B=Name(filename, hyperlink), C=Title, D=Modified,
        #          E=Description, F=Archived, G=Owner, H=Org, I=ItemType,
        #          J=Path, K=ProductFamily
        if len(row) < 11:
            continue
        name_cell = row[1]
        title_cell = row[2]
        desc_cell = row[4]
        pf_cell = row[10]
        filename = s(name_cell.value)
        title = s(title_cell.value)
        doc_name = title or filename
        if not doc_name:
            continue
        url = name_cell.hyperlink.target if name_cell.hyperlink else ''
        desc = s(desc_cell.value)
        pf_raw = s(pf_cell.value)
        portfolio = []
        if pf_raw:
            for p in pf_raw.split(','):
                p = NORMALIZE.get(p.strip(), p.strip())
                if p in VALID_PORTFOLIO and p not in portfolio:
                    portfolio.append(p)
        tags = assign_tags(f"{doc_name} {filename}", desc)
        rows.append({
            'name': doc_name,
            'url': url,
            'description': desc,
            'portfolio': portfolio,
            'tags': tags,
            'important': False,
            'lastEdited': '',
        })
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True, help='Path to OneStore xlsx')
    ap.add_argument('--extras', default='', help='Optional JSON with hand-curated extras + overrides')
    ap.add_argument('--out', default='notion-documents.json', help='Output JSON path')
    args = ap.parse_args()

    docs = read_xlsx(args.xlsx)
    print(f'Read {len(docs)} rows from xlsx')

    # Optional overlay: { "extras": [...new docs...], "important": ["Name 1", "Name 2"] }
    if args.extras and os.path.exists(args.extras):
        with open(args.extras, encoding='utf-8') as f:
            cfg = json.load(f)
        important_set = set(cfg.get('important', []))
        extras = cfg.get('extras', [])
        # Mark Important by exact name match
        for d in docs:
            if d['name'] in important_set:
                d['important'] = True
        # Append extras (these can also have important=true)
        for e in extras:
            docs.append(e)
        print(f'Applied {len(important_set)} important flags, appended {len(extras)} extras')

    # Sort: important first, then alphabetical
    docs.sort(key=lambda d: (0 if d['important'] else 1, d['name'].lower()))

    today = datetime.now(timezone(timedelta(hours=7))).strftime('%Y-%m-%d')
    out = {
        'lastSync': today,
        'source': 'OneStore xlsx + manual config (offline-generated)',
        'count': len(docs),
        'documents': docs,
    }
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
        f.write('\n')
    print(f'Wrote {len(docs)} docs to {args.out}')


if __name__ == '__main__':
    main()
